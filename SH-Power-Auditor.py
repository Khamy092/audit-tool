import requests
import ssl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import json
import time
from openpyxl import load_workbook
import ttk
from tkinter.filedialog import askopenfilename
from tkinter import *
from ttk import *
from tkinter.ttk import *
import os, sys
from pyunpack import Archive
from sys import exit

listDirectory = ''
downloadFlag = 0
username = ''
password = ''
usernameInput = ''
passwordInput = ''
downloadButton = ''


class Window(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.master = master
        self.init_window()
    
    
    def init_window(self):
        self.master.title("Power Auditor")
        self.pack(fill=BOTH, expand=1)
        quitButton = Button(self, text="Quit", command=self.quitGUI)
        quitButton.place(x=300, y=250)
        Title = Label(self, text="SiteHandler Power Auditor")
        Title.place(relx=0.5, rely=0.05, anchor=CENTER)

        def getPath():
            global listDirectory
            listDirectory = askopenfilename(filetypes=[('Excel file','.xlsx')])


        def download():
            progress.start()
            root.update()
            global listDirectory
            global usernameInput, passwordInput, username, password
            text = Label(self, text='Loading SiteHandler...')
            text.place(relx=0.90, rely=0.57, anchor=E)
            root.update()

            username = usernameInput.get()
            password = passwordInput.get()
            #print("username: " + username)
            #print("password: " + password)
            #print("listDirectory: " + str(listDirectory))
            workbook1 = load_workbook(str(listDirectory))
            siteList = workbook1.worksheets[0]
            workbook2 = load_workbook('./sites-with-ID.xlsx')
            objectIDList = workbook2.worksheets[0]
            driver = webdriver.Chrome(executable_path='./chromedriver')

            driver.get("https://sitehandler-anza.internal.ericsson.com/sh-anza/login;")
            assert "IS Tools" in driver.title

            element = driver.find_element_by_id("j_username")
            element.send_keys(username)
            element = driver.find_element_by_id("j_password")
            element.send_keys(password)
            element.send_keys(Keys.RETURN)
            #print("logged in")

            currentURL = driver.current_url

            while ("killsessions" in currentURL or "selectapplication" in currentURL or "sitehandler-anza.internal.ericsson.com/sh-anza/login" in currentURL):
                time.sleep(3)
                currentURL = driver.current_url


            cookie = driver.current_url
            cookieIndex = cookie.find('cookie')
            cookie = cookie[cookieIndex:]
            #print(cookie)

            x = 1
            y = 1
            sheet1Column1 = 'A' + str(x)
            sheet1Column2 = 'B' + str(x)
            sheet2Column1 = 'A' + str(y)
            sheet2Column2 = 'B' + str(y)
            sheet2Column3 = 'C' + str(y)

            text.config(text='Downloading Files...')
            while siteList[sheet1Column1].value != None:
                progress.step(1)
                root.update()
                sheet1Column1 = 'A' + str(x)
                sheet1Column2 = 'B' + str(x)
                #x = 1
                site1 = siteList[sheet1Column1].value
                #print(site1)
                found = 0
                while objectIDList[sheet2Column1].value != None and site1 != None:
                    sheet2Column1 = 'A' + str(y)
                    sheet2Column2 = 'B' + str(y)
                    sheet2Column3 = 'C' + str(y)
                    objectID = objectIDList[sheet2Column1].value
                    site2 = objectIDList[sheet2Column2].value
                    if site1 in site2:
                        found = 1
                        siteList[sheet1Column2] = objectIDList[sheet2Column3].value
                        downloadLink = "https://sitehandler-anza.internal.ericsson.com/sh-anza/legacy/FileSystemFileController?node=136&action=GETFILE&objId=" + str(objectID) + "&revision=0&propId=264128&" + cookie
                        #print(downloadLink)
                        driver.get(downloadLink)
                        fileName = os.listdir(os.path.expanduser('~')+"/Downloads/")
                        y = 1
                        break
                    else: 
                        y +=1
                        sheet2Column1 = 'A' + str(y)
                        sheet2Column2 = 'B' + str(y)
                        sheet2Column3 = 'C' + str(y)
                if found == 0:
                    y = 1
                    sheet2Column1 = 'A' + str(y)
                    sheet2Column2 = 'B' + str(y)
                    sheet2Column3 = 'C' + str(y)
                x+=1
            
            text.config(text='Download Complete!')
            text.place(relx=0.89, rely=0.57, anchor=E)
            root.update()
            progress.stop()
            time.sleep(3)
            driver.quit()
            root.update()
            workbook1.save(listDirectory)

        def powerAudit():
            global listDirectory
            if listDirectory != '':
                text = Label(self, text='Running Audit...')
                text.place(relx=0.5, rely=0.83, anchor=CENTER)
                root.update()
                powerCalcFiles = os.listdir(os.path.expanduser('~')+"/Downloads/")
                workbook1 = load_workbook(str(listDirectory))
                siteList = workbook1.worksheets[0]
                x = 1
                sheet1ColumnA = 'A' + str(x)
                sheet1ColumnB = 'B' + str(x)
                sheet1ColumnC = 'C' + str(x)
                sheet1ColumnD = 'D' + str(x)
                while siteList[sheet1ColumnA].value != None:
                    index2 = 24
                    batteryBackupCell = 'B' + str(index2)
                    if siteList[sheet1ColumnB].value != None:
                        for index, value in enumerate(powerCalcFiles):
                            underscoreLocation = value.find('_')
                            siteInFileName = value[:underscoreLocation]
                            if siteInFileName[-1] == ' ':
                                siteInFileName = siteInFileName[:underscoreLocation-1]
                                print(siteInFileName)

                            if siteInFileName == siteList[sheet1ColumnB].value:
                                print(siteInFileName)
                                #and '.zip' not in value and '.rar' not in value
                                siteList[sheet1ColumnC] = powerCalcFiles[index]
                                downloadFile = os.path.expanduser('~')+"/Downloads/"+value
                                print(downloadFile)
                                if '.zip' not in value and '.rar' not in value:
                                    workbook3 = load_workbook(downloadFile)

                                    if 'Battery Dimensioning' in workbook3.sheetnames:
                                        batteryDimensioning = workbook3['Battery Dimensioning']
                                    while batteryDimensioning['B'+str(index2+1)].value:
                                        index2 +=1
                                        print(x)
                                        print('-> next row')

                                    batteryBackupCell = 'B' + str(index2)
                                    siteList[sheet1ColumnD] = batteryDimensioning[batteryBackupCell].value
                                else: 
                                    siteList[sheet1ColumnD] = '.zip or .rar file requires manual extraction'
                                print(value)
                                break
                            elif index == len(powerCalcFiles)-1:
                                siteList[sheet1ColumnC] = 'No power calculation downloaded'
                                siteList[sheet1ColumnD] = 'No power calculation downloaded'
                    else:
                        siteList[sheet1ColumnB] = 'Not Found'
                        siteList[sheet1ColumnC] = 'Not Found'
                        siteList[sheet1ColumnD] = 'Not Found'
                    x +=1
                    sheet1ColumnA = 'A' + str(x)
                    sheet1ColumnB = 'B' + str(x)
                    sheet1ColumnC = 'C' + str(x)
                    sheet1ColumnD = 'D' + str(x)

                siteList.insert_rows(1)
                siteList['A1'] = 'POW'
                siteList['B1'] = 'Site Name'
                siteList['C1'] = 'File Name'
                siteList['D1'] = 'Battery Backup Time (h)'
                workbook1.save(listDirectory)

                text.config(text='Audit Complete!')
                text2 = Label(self, text='Check your spreadsheet')
                text2.place(relx=0.5, rely=0.89, anchor=CENTER)
                root.update()






        entryText1 = Label(self, text="SiteHandler Power Auditor", font=(16))
        entryText1.pack()
        entryText2 = Label(self, text="1. Browse the \".xlsx\" spreadsheet containing the list of sites:")
        entryText2.place(relx=0.80, rely=0.15, anchor=E)
        uploadButton = Button(self, text="Browse", command=getPath)
        uploadButton.place(relx=0.50, rely=0.25, anchor=CENTER)
        entryText3 = Label(self, text="2. Log into SiteHandler to download the power calculations:")
        global usernameInput
        entryText3.place(relx=0.81, rely=0.35, anchor=E)
        entryText4 = Label(self, text="Username: ")
        entryText4.place(relx=0.20, rely=0.45, anchor=E)
        usernameInput = Entry(self, width=15)
        usernameInput.place(relx=0.45, rely=0.45, anchor=E)
        global passwordInput
        entryText5 = Label(self, text="Password: ")
        entryText5.place(relx=0.65, rely=0.45, anchor=E)
        passwordInput = Entry(self, width=15,  show="*",)
        passwordInput.place(relx=0.90, rely=0.45, anchor=E)
        #passwordInput.trace_add('write', toggleButton)
        global downloadButton
        downloadButton = Button(self, text="Download", command=download)
        downloadButton.place(relx=0.50, rely=0.57, anchor=CENTER)
        entryText6 = Label(self, text="3. Run Power Audit:")
        entryText6.place(relx=0.28, rely=0.65, anchor=E)
        auditButton = Button(self, text="Run Power Audit", command=powerAudit)
        auditButton.place(relx=0.50, rely=0.75, anchor=CENTER)
        Footer = Label(self, text='Updated Oct 2019 (eqasali)')
        Footer.place(relx=0.2, rely=0.95, anchor=CENTER)

        

    ssl._create_default_https_context = ssl._create_unverified_context


    def quitGUI(self):
        exit()



root = Tk()
root.geometry("400x300")
theme = Style()
theme.theme_use('vista')
#print(theme.theme_names())
root.resizable(False, False)
progress = Progressbar(root, orient=HORIZONTAL, length=120)
progress.place(x=0, y=0)
#progress.place(relx=0.01, rely=0.01, anchor=CENTER)
progress.config(mode = 'indeterminate', maximum=100, value=1)
app = Window(root)
root.mainloop()

